"""
Check what's in the KB25019 Contract sheet
"""

import sys
import os
sys.path.append(os.path.dirname(__file__))

import openpyxl


def check_kb25019_contract():
    """Check what's in the KB25019 Contract sheet."""
    
    excel_file = "CT&INV&PL KB25019 DAP(1)(1).xlsx"
    
    if not os.path.exists(excel_file):
        print("❌ KB25019 Excel file not found")
        return
    
    print(f"🔍 Checking KB25019 Contract Sheet")
    print("="*60)
    
    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=False)
        
        if 'Contract' not in workbook.sheetnames:
            print("❌ No Contract sheet found")
            print(f"Available sheets: {workbook.sheetnames}")
            return
        
        worksheet = workbook['Contract']
        
        print(f"📋 CONTRACT SHEET CONTENT:")
        
        # Check table area (rows 15-25)
        print(f"\n📊 TABLE AREA (Rows 15-25):")
        for row in range(15, 26):
            row_data = []
            for col in range(1, 15):  # Columns A-N
                cell = worksheet.cell(row, col)
                if cell.value:
                    text = str(cell.value).strip()[:30]  # First 30 chars
                    if text:
                        row_data.append(f"Col{col}:'{text}'")
            if row_data:
                print(f"   Row {row}: {', '.join(row_data)}")
        
        # Check merged cells in Contract sheet
        print(f"\n🔗 MERGED CELLS:")
        merged_ranges = list(worksheet.merged_cells.ranges)
        contract_merged = [r for r in merged_ranges if 15 <= r.min_row <= 25]
        
        if contract_merged:
            for i, merged_range in enumerate(contract_merged, 1):
                top_left = worksheet.cell(merged_range.min_row, merged_range.min_col)
                text = str(top_left.value).strip() if top_left.value else ""
                rowspan = merged_range.max_row - merged_range.min_row + 1
                colspan = merged_range.max_col - merged_range.min_col + 1
                print(f"   [{i}] Row {merged_range.min_row}: '{text}' ({rowspan}×{colspan})")
        else:
            print("   No merged cells in table area")
        
        # Also check earlier rows in case table is higher up
        print(f"\n📊 EARLIER ROWS (Rows 10-14):")
        for row in range(10, 15):
            row_data = []
            for col in range(1, 15):
                cell = worksheet.cell(row, col)
                if cell.value:
                    text = str(cell.value).strip()[:30]
                    if text:
                        row_data.append(f"Col{col}:'{text}'")
            if row_data:
                print(f"   Row {row}: {', '.join(row_data)}")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    check_kb25019_contract()
