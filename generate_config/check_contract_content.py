"""
Check what's in the Contract sheet of KB25019
"""

import openpyxl


def check_contract_sheet():
    """Check the Contract sheet content in KB25019."""
    
    excel_file = "../CT&INV&PL KB25019 DAP(1)(1).xlsx"
    
    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=False)
        worksheet = workbook['Contract']
        
        print(f"🔍 Contract Sheet Content (rows 15-25):")
        
        for row in range(15, 26):
            row_data = []
            for col in range(1, 15):  # Columns A-N
                cell = worksheet.cell(row, col)
                if cell.value:
                    text = str(cell.value).strip()[:20]  # First 20 chars
                    if text:
                        row_data.append(f"Col{col}:'{text}'")
            if row_data:
                print(f"   Row {row}: {', '.join(row_data)}")
        
        print(f"\n🔍 Contract Sheet Merged Cells in rows 15-25:")
        merged_ranges = list(worksheet.merged_cells.ranges)
        table_merges = [r for r in merged_ranges if 15 <= r.min_row <= 25]
        
        for merged_range in table_merges:
            top_left = worksheet.cell(merged_range.min_row, merged_range.min_col)
            text = str(top_left.value).strip() if top_left.value else ""
            rowspan = merged_range.max_row - merged_range.min_row + 1
            colspan = merged_range.max_col - merged_range.min_col + 1
            print(f"   Row {merged_range.min_row}: '{text}' ({rowspan}×{colspan})")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")


if __name__ == "__main__":
    check_contract_sheet()
