#!/usr/bin/env python3
"""
Height checker for JLFMH Excel file - compare actual vs config heights
"""

from openpyxl import load_workbook
import json

def check_header_heights():
    # Load the Excel file
    wb = load_workbook('CT&INV&PL JLFMH25021 DAP(2).xlsx')
    
    print('=== JLFMH EXCEL FILE - ALL HEADER HEIGHTS ===\n')
    
    # Check Invoice
    if 'Invoice' in wb.sheetnames:
        sheet = wb['Invoice']
        row = 21
        height = sheet.row_dimensions[row].height
        print(f'📋 Invoice Sheet:')
        print(f'   Header Row: {row}')
        print(f'   Actual Height: {height}pt')
        print(f'   Header Content: ["Mark & Nº", "P.O. Nº", "ITEM Nº", "Description", "Quantity(SF)", "Unit price (USD)", "Amount (USD)"]')
    
    # Check Packing list  
    if 'Packing list' in wb.sheetnames:
        sheet = wb['Packing list']
        print(f'\n📋 Packing list Sheet:')
        
        # Find header row
        header_row = None
        for row in range(1, 30):
            cell_values = []
            for col in range(1, 12):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    cell_values.append(str(cell.value).strip())
            
            row_text = ' '.join(cell_values)
            if any(keyword in row_text.upper() for keyword in ['PALLET', 'ITEM', 'DESCRIPTION', 'QUANTITY', 'PCS', 'SF']):
                header_row = row
                height = sheet.row_dimensions[row].height
                print(f'   Header Row: {row}')
                print(f'   Actual Height: {height}pt')
                print(f'   Header Content: {cell_values[:6]}...')
                break
    
    # Load the generated config file
    config_path = 'result/CT&INV&PL JLFMH25021 DAP(2)/CT&INV&PL JLFMH25021 DAP(2)_config.json'
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    print('\n=== CONFIG FILE COMPARISON ===\n')
    
    # Find config heights - search through the entire config structure
    def find_row_heights(data, path=""):
        """Recursively find row_heights sections in config"""
        results = []
        if isinstance(data, dict):
            for key, value in data.items():
                current_path = f"{path}.{key}" if path else key
                if key == "row_heights" and isinstance(value, dict):
                    results.append((current_path, value))
                else:
                    results.extend(find_row_heights(value, current_path))
        elif isinstance(data, list):
            for i, item in enumerate(data):
                current_path = f"{path}[{i}]"
                results.extend(find_row_heights(item, current_path))
        return results
    
    # Find all row_heights sections
    row_heights_sections = find_row_heights(config)
    
    print(f'📊 Found {len(row_heights_sections)} row_heights sections in config:')
    for i, (path, heights) in enumerate(row_heights_sections, 1):
        print(f'   {i}. {path}: {heights}')
    
    # Compare with actual Excel heights
    print(f'\n📊 Header Height Comparison:')
    print(f'   Invoice Excel: 45.0pt')
    
    if row_heights_sections:
        # Assume first section is Invoice (as per config structure)
        invoice_config_height = row_heights_sections[0][1].get('header', 'Not found')
        print(f'   Invoice Config: {invoice_config_height}pt')
        
        if invoice_config_height and invoice_config_height != 'Not found':
            diff = 45.0 - float(invoice_config_height)
            status = '✅ MATCH' if abs(diff) <= 1 else '❌ MISMATCH'
            print(f'   Difference: {diff}pt - {status}')
        else:
            print(f'   Status: ❌ CONFIG HEIGHT NOT FOUND')
    
    return row_heights_sections

if __name__ == "__main__":
    check_header_heights()
